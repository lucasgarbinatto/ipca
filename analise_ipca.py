import json
from collections import defaultdict
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- Criar pasta de relatórios ---
PASTA_RELATORIOS = 'relatórios'
if not os.path.exists(PASTA_RELATORIOS):
    os.makedirs(PASTA_RELATORIOS)

# --- Carregar dados ---
with open("expectativas_filtradas.json", "r", encoding="utf-8") as f:
    dados = json.load(f)

# --- Pegar apenas a data de pesquisa mais recente ---
data_mais_recente = max(d["Data"] for d in dados)
dados_filtrados = [d for d in dados if d["Data"] == data_mais_recente]

print(f"=" * 60)
print(f"ANÁLISE DAS EXPECTATIVAS DE IPCA - TOP 5 BCB")
print(f"Data da pesquisa: {data_mais_recente}")
print(f"=" * 60)

# --- Organizar por mês de referência ---
expectativas = {}
for d in dados_filtrados:
    ref = d["DataReferencia"]  # formato: "MM/YYYY"
    expectativas[ref] = {
        "mediana": d["Mediana"],
        "media": d["Media"],
        "minimo": d["Minimo"],
        "maximo": d["Maximo"],
        "desvio": d["DesvioPadrao"]
    }

# --- Função para calcular inflação acumulada (composta) ---
def calcular_inflacao_acumulada(taxas_mensais):
    """
    Calcula inflação acumulada usando fórmula composta:
    (1 + r1/100) * (1 + r2/100) * ... * (1 + rn/100) - 1
    Retorna em percentual
    """
    acumulado = 1.0
    for taxa in taxas_mensais:
        acumulado *= (1 + (taxa / 100))
    return (acumulado - 1) * 100

# --- Definir meses por ano ---
meses_2026 = [f"{m:02d}/2026" for m in range(1, 13)]
meses_2027 = [f"{m:02d}/2027" for m in range(1, 13)]

# --- Função para exibir análise de um ano ---
def analisar_ano(ano, meses):
    print(f"\n{'=' * 60}")
    print(f"EXPECTATIVAS MENSAIS DE IPCA - {ano}")
    print(f"{'=' * 60}")
    print(f"{'Mês':<12} {'Mediana':>10} {'Mínimo':>10} {'Máximo':>10} {'Desvio':>10}")
    print("-" * 60)
    
    taxas_mensais = []
    trimestres = defaultdict(list)
    
    for i, mes in enumerate(meses, 1):
        if mes in expectativas:
            exp = expectativas[mes]
            taxas_mensais.append(exp["mediana"])
            
            # Determinar trimestre (1-3: Q1, 4-6: Q2, 7-9: Q3, 10-12: Q4)
            trimestre = (i - 1) // 3 + 1
            trimestres[trimestre].append(exp["mediana"])
            
            print(f"{mes:<12} {exp['mediana']:>10.4f}% {exp['minimo']:>10.4f}% {exp['maximo']:>10.4f}% {exp['desvio']:>10.4f}")
        else:
            print(f"{mes:<12} {'N/D':>10} {'N/D':>10} {'N/D':>10} {'N/D':>10}")
    
    # --- Inflação acumulada 12 meses ---
    if taxas_mensais:
        inflacao_12m = calcular_inflacao_acumulada(taxas_mensais)
        soma_simples = sum(taxas_mensais)
        
        print("-" * 60)
        print(f"\n📊 INFLAÇÃO ACUMULADA {ano} (12 meses)")
        print(f"   Acumulada (composta): {inflacao_12m:.2f}%")
        print(f"   Soma simples:         {soma_simples:.2f}%")
        
        # Identificar meses críticos
        if taxas_mensais:
            max_taxa = max(taxas_mensais)
            min_taxa = min(taxas_mensais)
            mes_max = meses[taxas_mensais.index(max_taxa)]
            mes_min = meses[taxas_mensais.index(min_taxa)]
            media_mensal = sum(taxas_mensais) / len(taxas_mensais)
            
            print(f"\n📈 INSIGHTS {ano}:")
            print(f"   Maior expectativa:  {max_taxa:.4f}% em {mes_max}")
            print(f"   Menor expectativa:  {min_taxa:.4f}% em {mes_min}")
            print(f"   Média mensal:       {media_mensal:.4f}%")
    
    # --- Inflação por trimestre ---
    print(f"\n📅 INFLAÇÃO POR TRIMESTRE {ano}")
    print(f"{'Trimestre':<15} {'Meses':<20} {'Acumulado':>12}")
    print("-" * 50)
    
    nomes_trimestres = {
        1: "Q1 (Jan-Mar)",
        2: "Q2 (Abr-Jun)",
        3: "Q3 (Jul-Set)",
        4: "Q4 (Out-Dez)"
    }
    
    for q in range(1, 5):
        if q in trimestres and trimestres[q]:
            acum_tri = calcular_inflacao_acumulada(trimestres[q])
            n_meses = len(trimestres[q])
            print(f"{nomes_trimestres[q]:<15} {n_meses} meses{'':<13} {acum_tri:>10.2f}%")
        else:
            print(f"{nomes_trimestres[q]:<15} {'N/D':<20} {'N/D':>12}")
    
    return taxas_mensais, trimestres

# --- Executar análise para 2026 e 2027 ---
taxas_2026, tri_2026 = analisar_ano(2026, meses_2026)
taxas_2027, tri_2027 = analisar_ano(2027, meses_2027)

# --- Comparativo entre anos ---
print(f"\n{'=' * 60}")
print("📊 COMPARATIVO 2026 vs 2027")
print(f"{'=' * 60}")

if taxas_2026 and taxas_2027:
    inf_2026 = calcular_inflacao_acumulada(taxas_2026)
    inf_2027 = calcular_inflacao_acumulada(taxas_2027)
    diferenca = inf_2027 - inf_2026
    
    print(f"   Inflação acumulada 2026: {inf_2026:.2f}%")
    print(f"   Inflação acumulada 2027: {inf_2027:.2f}%")
    print(f"   Diferença (2027-2026):   {diferenca:+.2f}%")
    
    if diferenca > 0:
        print(f"\n   ⚠️  Expectativa de AUMENTO da inflação em 2027")
    elif diferenca < 0:
        print(f"\n   ✅ Expectativa de REDUÇÃO da inflação em 2027")
    else:
        print(f"\n   ➡️  Expectativa de inflação ESTÁVEL")

# --- Meta de inflação ---
print(f"\n{'=' * 60}")
print("🎯 COMPARAÇÃO COM META DE INFLAÇÃO")
print(f"{'=' * 60}")
META_INFLACAO = 3.0  # Meta do BCB
TETO_META = 4.5      # Teto (meta + 1.5pp)

if taxas_2026:
    inf_2026 = calcular_inflacao_acumulada(taxas_2026)
    desvio_meta = inf_2026 - META_INFLACAO
    
    print(f"   Meta de inflação:     {META_INFLACAO:.1f}%")
    print(f"   Teto da meta:         {TETO_META:.1f}%")
    print(f"   Expectativa 2026:     {inf_2026:.2f}%")
    print(f"   Desvio da meta:       {desvio_meta:+.2f}pp")
    
    if inf_2026 <= META_INFLACAO:
        print(f"\n   ✅ Expectativa DENTRO da meta")
    elif inf_2026 <= TETO_META:
        print(f"\n   ⚠️  Expectativa ACIMA da meta, mas dentro do teto")
    else:
        print(f"\n   🚨 Expectativa ACIMA do teto da meta!")

print(f"\n{'=' * 60}")
print("Análise concluída!")
print(f"{'=' * 60}")

# --- Exportar para Excel ---
def exportar_para_excel(expectativas, taxas_2026, taxas_2027, meses_2026, meses_2027, data_pesquisa):
    """Exporta os dados de expectativas para um arquivo Excel formatado."""
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # --- Aba 1: Expectativas Mensais ---
    ws1 = wb.active
    ws1.title = "Expectativas Mensais"
    
    # Título
    ws1['A1'] = f"Expectativas IPCA - Top 5 BCB (Pesquisa: {data_pesquisa})"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:F1')
    
    # Cabeçalhos
    headers = ["Mês/Ano", "Mediana (%)", "Média (%)", "Mínimo (%)", "Máximo (%)", "Desvio Padrão"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Dados 2026
    row = 4
    ws1.cell(row=row, column=1, value="--- 2026 ---").font = Font(bold=True)
    row += 1
    
    for mes in meses_2026:
        if mes in expectativas:
            exp = expectativas[mes]
            ws1.cell(row=row, column=1, value=mes).border = border
            ws1.cell(row=row, column=2, value=exp["mediana"]).border = border
            ws1.cell(row=row, column=3, value=exp["media"]).border = border
            ws1.cell(row=row, column=4, value=exp["minimo"]).border = border
            ws1.cell(row=row, column=5, value=exp["maximo"]).border = border
            ws1.cell(row=row, column=6, value=exp["desvio"]).border = border
            for col in range(1, 7):
                ws1.cell(row=row, column=col).alignment = center_align
            row += 1
    
    # Dados 2027
    row += 1
    ws1.cell(row=row, column=1, value="--- 2027 ---").font = Font(bold=True)
    row += 1
    
    for mes in meses_2027:
        if mes in expectativas:
            exp = expectativas[mes]
            ws1.cell(row=row, column=1, value=mes).border = border
            ws1.cell(row=row, column=2, value=exp["mediana"]).border = border
            ws1.cell(row=row, column=3, value=exp["media"]).border = border
            ws1.cell(row=row, column=4, value=exp["minimo"]).border = border
            ws1.cell(row=row, column=5, value=exp["maximo"]).border = border
            ws1.cell(row=row, column=6, value=exp["desvio"]).border = border
            for col in range(1, 7):
                ws1.cell(row=row, column=col).alignment = center_align
            row += 1
    
    # Ajustar largura das colunas
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    
    # --- Aba 2: Resumo Anual ---
    ws2 = wb.create_sheet(title="Resumo Anual")
    
    ws2['A1'] = "Resumo de Inflação Acumulada"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')
    
    # Cabeçalhos
    for col, header in enumerate(["Ano", "Inflação Acumulada (%)", "Média Mensal (%)"], 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Dados
    if taxas_2026:
        inf_2026 = calcular_inflacao_acumulada(taxas_2026)
        media_2026 = sum(taxas_2026) / len(taxas_2026)
        ws2.cell(row=4, column=1, value=2026).border = border
        ws2.cell(row=4, column=2, value=round(inf_2026, 4)).border = border
        ws2.cell(row=4, column=3, value=round(media_2026, 4)).border = border
    
    if taxas_2027:
        inf_2027 = calcular_inflacao_acumulada(taxas_2027)
        media_2027 = sum(taxas_2027) / len(taxas_2027)
        ws2.cell(row=5, column=1, value=2027).border = border
        ws2.cell(row=5, column=2, value=round(inf_2027, 4)).border = border
        ws2.cell(row=5, column=3, value=round(media_2027, 4)).border = border
    
    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 22
        for row in range(4, 6):
            ws2.cell(row=row, column=col).alignment = center_align
    
    # --- Aba 3: Comparação com Meta ---
    ws3 = wb.create_sheet(title="Meta de Inflação")
    
    ws3['A1'] = "Comparação com Meta de Inflação BCB"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:C1')
    
    ws3.cell(row=3, column=1, value="Meta de Inflação:").font = Font(bold=True)
    ws3.cell(row=3, column=2, value="3.0%")
    ws3.cell(row=4, column=1, value="Teto da Meta:").font = Font(bold=True)
    ws3.cell(row=4, column=2, value="4.5%")
    
    if taxas_2026:
        inf_2026 = calcular_inflacao_acumulada(taxas_2026)
        desvio = inf_2026 - 3.0
        ws3.cell(row=6, column=1, value="Expectativa 2026:").font = Font(bold=True)
        ws3.cell(row=6, column=2, value=f"{inf_2026:.2f}%")
        ws3.cell(row=7, column=1, value="Desvio da Meta:").font = Font(bold=True)
        ws3.cell(row=7, column=2, value=f"{desvio:+.2f}pp")
        
        if inf_2026 <= 3.0:
            status = "✅ Dentro da meta"
        elif inf_2026 <= 4.5:
            status = "⚠️ Acima da meta, dentro do teto"
        else:
            status = "🚨 Acima do teto!"
        ws3.cell(row=8, column=1, value="Status:").font = Font(bold=True)
        ws3.cell(row=8, column=2, value=status)
    
    for col in range(1, 3):
        ws3.column_dimensions[get_column_letter(col)].width = 25
    
    # Salvar arquivo
    nome_arquivo = f"expectativas_ipca_{data_pesquisa.replace('/', '-')}.xlsx"
    wb.save(nome_arquivo)
    print(f"\n📁 Dados exportados para: {nome_arquivo}")
    return nome_arquivo

# Executar exportação
exportar_para_excel(expectativas, taxas_2026, taxas_2027, meses_2026, meses_2027, data_mais_recente)
